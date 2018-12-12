# -*- coding:utf-8 -*-

__author__ = 'cutieyy'

import random
import os
import xlsxwriter
from openpyxl import Workbook

'''
if __name__=="__main__":
    print("写Excel文件简单实例")

    wb=Workbook()

    ws=wb.active

    ws['A1']="数字编号"
    ws['B1']="生成的随机数字（1-1000）"

    for i in range(1,101):
        ws.append(i,random.randint(1,1000))


    wb.save("random.xlsx")

'''


class XlsAPI:
    # Excel报告接口
    def __init__(self, file_path):
        self.xls_workbook = None  # 创建Excel对象
        self.xls_file_path = file_path  # 创建Excel的文件路径
        self.xls_worksheets = {}  # 工作表的行列标记符{worksheet对象：{r行，c列}}
        self.xls_sheet_col_length = {}  # 工作表的列宽{worksheet对象：{列1：宽度，列2：宽度..}}

    def xlsOpenWorkbook(self):
        # 创建一个Excel
        self.xls_workbook = xlsxwriter.Workbook(self.xls_file_path)

    def xlsAddWorksheet(self, sheet_name='sheet', r=0, c=0):
        # 创建一个工作簿对象
        obj_worksheet = self.xls_workbook.add_worksheet(sheet_name)  # 创建工作簿对象
        xls_worksheet = {}  # 工作簿的行列标记符
        xls_worksheet['r'] = r  # 行标记符
        xls_worksheet['c'] = c  # 列标记符
        self.xls_worksheets[obj_worksheet] = xls_worksheet
        self.xls_sheet_col_length[obj_worksheet] = {}
        return obj_worksheet

    def xlsCloseWorkbook(self):
        # 关闭Excel文件对象，保存Excel数据
        # 关闭工作簿对象
        self.xls_workbook.close()

    def addWorksheetTitle(self, obj_worksheet, titles=[], r=0, c=0):
        # 添加工作簿的标题
        # 设置标题样式
        style_title = self.xls_workbook.add_format({'bold': True})
        style_title.set_pattern(1)  # 填充颜色
        style_title.set_bg_color('green')

        # 写工作簿的标题
        for title in titles:
            obj_worksheet.write(r, c, title, style_title)
            c += 1

         # 工作簿的行标记+1

        r += 1
        self.xls_worksheets[obj_worksheet]['r'] = r

    def appendWorkshetData(self, obj_worksheet, datas=[], r=None, c=None, gold=0):
        # 按行追加数据

        # 标准
        stylebox = self.xls_workbook.add_format()
        # 红体
        stylebox_red = self.xls_workbook.add_format({'font_color': 'red'})
        # 蓝体
        stylebox_blue = self.xls_workbook.add_format({'font_color': 'blue'})
        # 写工作簿的行
        if None == r:
            r = self.xls_worksheets[obj_worksheet]['r']
        else:
            r = r
        # 写工作簿的列
        if None == c:
            c = self.xls_worksheets[obj_worksheet]['c']
        else:
            c = c
        # 写工作簿的数据
        for data in datas:
            if str != type(data):
                data = str(data)
            # parameter gold:0(不变/pass),-1(变差/fail)，1（变好）
            if 0 == gold:
                obj_worksheet.write(r, c, data, stylebox)
            elif -1 == gold:
                obj_worksheet.write(r, c, data, stylebox_red)
            elif 1 == gold:
                obj_worksheet.write(r, c, data, stylebox_blue)
            else:
                print('info:gold was illegal.')
                obj_worksheet.write(r, c, data, stylebox)

            # 工作簿每列的最大字符长度
            if c not in self.xls_sheet_col_length[obj_worksheet]:
                self.xls_sheet_col_length[obj_worksheet][c] = len(data)
            else:
                if self.xls_sheet_col_length[obj_worksheet][c] < len(data):
                    self.xls_sheet_col_length[obj_worksheet][c] = len(data)
            c += 1
        #工作簿的行标记+1
        r += 1
        self.xls_worksheets[obj_worksheet]['r'] = r

    def setWorksheetWidth(self, obj_worksheet):
        # 设置Excel表宽度
        list_col = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

        # 设置工作簿的列宽

        for c in self.xls_sheet_col_length[obj_worksheet]:
            width_string = ''.join([list_col[c],':',list_col[c]])
            if self.xls_sheet_col_length[obj_worksheet][c] < 10:
                obj_worksheet.set_column(width_string, 10)
            elif self.xls_sheet_col_length[obj_worksheet][c] > 50:
                obj_worksheet.set_column(width_string, 10)
            else:
                obj_worksheet.set_column(width_string, self.xls_sheet_col_length[obj_worksheet][c])


# self test


if __name__ == '__main__':
    file_name=input("请输入文件名：")
    file_path = os.path.join(os.getcwd(), file_name)
    xls = XlsAPI(file_path)
    # 创建Excel对象
    xls.xlsOpenWorkbook()
    # 添加工作对象
    sheet_names = ['sheet1', 'sheet2', 'sheet3']
    for sheet_name in sheet_names:
        sheet = xls.xlsAddWorksheet(sheet_name)
        # Excel的标题
        xls.addWorksheetTitle(sheet, ['数字编号', '生成的随机数（0-1000）'])
        # Excel的数据
        for i in range(1, 101):
            xls.appendWorkshetData(sheet, [i, random.randint(1, 1000)])

        xls.setWorksheetWidth(sheet)

    xls.xlsCloseWorkbook()
