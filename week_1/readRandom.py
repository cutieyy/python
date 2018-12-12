# -*- coding: utf-8 -*-
#读写Excel文件
import openpyxl
import xlwt
import xlrd

def readExcel(path):
    wb=openpyxl.load_workbook(path)
    sheet=wb.get_sheet_by_name("sheet1")
    for row in sheet.rows:
        for cell in row:
            print(cell.value,"\t",end="")
        print()

def writeExcel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_names()
    worksheet=wb.get_sheet_by_name(sheet[0])
    worksheet['C1']="乘积结果"
    for i in range(2,102):
        worksheet.cell(i,3).value=int(worksheet.cell(i,1).value)*int(worksheet.cell(i,2).value)
    wb.save(path)
    print ("写入数据成功！")

#self test

file_name=input("请输入文件名：")
file=file_name

readExcel(file)
writeExcel(file)