# -*- coding: utf-8 -*-
#读写Excel文件
import openpyxl
import xlwt
import xlrd
import os,sys

def readFile(path):
    dirs = os.listdir(path)
    # 输出所有文件和文件夹
    for file in dirs:
        print(file)
    print("文件夹读取成功！")

def readExcel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name("sheet1")
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

def writeExcel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_names()
    worksheet = wb.get_sheet_by_name(sheet[0])
    worksheet['C1'] = "乘积结果"
    j=0
    for i in range(2, 102):
        worksheet.cell(i, 3).value = int(worksheet.cell(i, 1).value) * int(worksheet.cell(i, 2).value)
        j+=1
    wb.save(path)
    print("写入数据成功！")
    return j

#self test


path0=r'/Users/momo/PycharmProjects/week_1/random_test'
readFile(path0)

str1="random"

for i in range(1,101):
    filename = str1 + str(i) + '.xlsx'
    file_path = os.path.join('/Users/momo/PycharmProjects/week_1/random_test', filename)
    readExcel(file_path)

count=0
for i in range(1,101):
    filename = str1 + str(i) + '.xlsx'
    file_path = os.path.join('/Users/momo/PycharmProjects/week_1/random_test', filename)
    writeExcel(file_path)
    count=writeExcel(file_path)+count

print(count)